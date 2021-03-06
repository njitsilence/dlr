# -*- coding: utf-8 -*-

import logging

from openpyxl import load_workbook

from vars_setting import PARMFILE_NAME, DEF_SHEET_NAME

#singleton mode via decorator 
def singleton(cls): 
	_instance = {}
	def _warper(*args, **kargs):
		if cls not in _instance:
			_instance[cls] = cls(*args, **kargs)
		return _instance[cls]
	return _warper

# Class definition

class RequestsLoader(object):
	"""docstring for ParameterLoader
	load the parameter file and retrun the list of request
	"""
	def __init__(self, parameter_filename=PARMFILE_NAME, sheetname=DEF_SHEET_NAME):
		self.parameter_filename = parameter_filename
		try:
			self.wb = load_workbook(self.parameter_filename)
		except Exception as e:
			raise

		try:
			self.sheet = self.wb.get_sheet_by_name(sheetname)
		except Exception as e:
			raise

	def get_records(self):
		records = []
		for row in range(1, self.sheet.max_row+1):
			records.append(self.sheet[row])
		return records

	def get_requests(self):
		records_list = []
		title_value = [col.value for col in self.sheet[1]]

		logging.debug('The maxrow of parameter file is %s' %self.sheet.max_row )
		
		for row in range(2, self.sheet.max_row+1):
			row_value = [ col.value for col in self.sheet[row]]
			records_list.append(dict(zip(title_value, row_value)))
		return records_list